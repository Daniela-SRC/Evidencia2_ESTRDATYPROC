import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import csv
import os
from os import path
import datetime
fecha_actual = datetime.date.today()
dia_actual = fecha_actual.day
mes_actual = fecha_actual.month
año_actual = fecha_actual.year
tupla_actual = (dia_actual, mes_actual, año_actual)
opcion = 0
clave_cliente = 0
clave_sala = 0
clave_registro = 0
fechaExistente = False
row=2

def menu():
    opc = int(input("Menú Principal\n" +
                    "Seleccione la opcion que guste:\n"+
                    "1.- Registrar la reservación de una sala\n" +
                    "2.- Editar el nombre de un evento reservado\n" +
                    "3.- Consultar reservaciones\n" +
                    "4.- Registrar un nuevo cliente\n" +
                    "5.- Registrar una sala\n" +
                    "6.- Consulta de Salas disponibles\n"+
                    "7.- Finalizar\n"))
    return opc


salas=[]
clientes=[]
eventos=[]
total_salas_turnos=[]
salas_turnos_ocupados=[]

while opcion !=7:
    opcion = menu()
    if opcion == 1:
        print("Registrar la reservación de una sala para un evento\n")
        if salas:
            clienteRegistrado = False            
            clave=int(input('Ingrese su ID: '))
            for elementoCliente in clientes:
                for validacionID in range(len(elementoCliente)):        
                    if clave == elementoCliente[0]:            
                        clienteRegistrado = True
                        break
                    else:                
                        break
            if clienteRegistrado:
                while True:
                    salaExistente = False
                    nombre_evento = input("Ingrese el nombre del evento: ") 
                    if nombre_evento != "": 
                        disponible = True 
                        cve_sala=int(input("Ingrese la clave de la sala del evento: ")) 
                        for revisionSala in salas:
                            for revisionClaveSala in range(len(revisionSala)):
                                if cve_sala == revisionSala[0]:
                                    salaExistente = True                                    
                                    break                                
                        if salaExistente:                    
                            for Lista in eventos: 
                                if disponible: 
                                    for claveIteracion in range(len(Lista)):                                                                                                                        
                                        if cve_sala == Lista[3]: 
                                            disponible = False
                                            break
                                else:
                                    break                                
                            if disponible: 
                                print("Continue con el registro")
                                while True:
                                    try:
                                        horario_evento = int(input("Ingrese el numero (1,2 ó 3) del horario del evento que desee (1.-MATUTINO, 2.-VESPERTINO, 3.-NOCTURNO): "))
                                        
                                    
                                    except ValueError:
                                        print("Formato de dato incorrecto")
                                    else:
                                        if horario_evento == 1:
                                            turno_a_guardar="Matutino"
                                            
                                        elif horario_evento == 2:
                                            turno_a_guardar="Vespertino"
                                            
                                        elif horario_evento == 3:
                                            turno_a_guardar="Nocturno"
                                        
                                        if horario_evento > 0 and horario_evento < 4:
                                            print("Horario guardado")
                                            break
                                        else:
                                            print("Horario no valido")
                                            
                                while True:
                                    fecha_reservada = input("Ingrese la fecha que desea reservar (dd/mm/aaaa): ")
                                    fecha_reservada = datetime.datetime.strptime(fecha_reservada,"%d/%m/%Y").date()
                                    dia_reservado = fecha_reservada.day
                                    mes_reservado = fecha_reservada.month
                                    año_reservado = fecha_reservada.year

                                    dia_valido = dia_reservado - dia_actual

                                    tupla_reservacion = (dia_reservado, mes_reservado, año_reservado)
                                    
                                    if dia_valido <= 1:
                                        print("Para reservar una fecha debe hacerlo con 2 dias de anticipación")
                                    else:
                                        if tupla_reservacion > tupla_actual:
                                            clave_registro += 1 
                                            print("Su reservación a sido éxitosa\n") 
                                            eventos.append((clave, clave_registro, nombre_evento, cve_sala, turno_a_guardar, tupla_reservacion))
                                            recorr=0
                                            for clave_sala, nombre_sala, cupo_sala in salas:                                    
                                                if cve_sala ==salas[recorr][0]:
                                                    salita=nombre_sala
                                                    salas_turnos_ocupados.append((salita, turno_a_guardar)) 
                                                recorr = recorr + 1
                                                
                                            break
                                        else:
                                            print("Para reservar una fecha debe hacerlo con 2 dias de anticipación")
                                break
                            else:
                                print("ERROR! La sala ya ha sido registrada")                                
                        else:
                            print("ERROR! No existe esa sala")                            
                    else:
                         break                                
            else:
                print("El cliente no está registrado")                                        
        else:
            print("ERROR! NO SE HA REGISTRADO ALGUNA SALA")
        
    if opcion == 2:
        print("Editar el nombre de un evento reservado\n")
        eventos = list(map(list,eventos))
        folio_evento = int(input("Folio del evento: "))
        
        for clave_evento in eventos:
            if clave_evento[0] == folio_evento:
                nombre_nuevo=input("Ingrese el nuevo nombre del evento: ")
                clave_evento[2]=nombre_nuevo
                print(eventos)
                
    if opcion == 3:
        print("Consultar reservaciones\n")
        fechaExistente=False
        fecha_consulta = input("Ingrese la fecha que desea consultar (dd/mm/aaaa): ")
        fecha_consulta = datetime.datetime.strptime(fecha_consulta,"%d/%m/%Y").date()
        dia_consulta = fecha_consulta.day
        mes_consulta = fecha_consulta.month
        año_consulta = fecha_consulta.year            
        tupla_consulta = (dia_consulta, mes_consulta, año_consulta)
        print("--------------------------------------------------------------------")
        print(f"**\t\tREPORTE DE RESERVACIONES PARA EL DIA {fecha_consulta}\t\t**")
        print("--------------------------------------------------------------------")
        print("SALA\t CLIENTE\t\t EVENTO\t\t TURNO")        
          
        libro = Workbook()
        hoja = libro.active
        hoja["A1"].value = "REPORTE DE EVENTOS PARA EL DIA: "
        hoja["B1"].value = fecha_consulta
        hoja["A2"].value= "SALA"
        hoja["B2"].value = "NOMBRE CLIENTE"
        hoja["C2"].value = "APELLIDO CLIENTE"
        hoja["D2"].value = "EVENTO"
        hoja["E2"].value = "TURNO"
        
        for num_cliente_buscar, id_buscar, nombre_buscar, sala_buscar, turno_buscar, fecha_buscar in eventos:            
            if tupla_consulta == fecha_buscar:                
                row=row+1
                for numero_sala, sala_nombre, sala_cupo in salas:
                    if numero_sala==sala_buscar:                        
                        imprimir_sala=numero_sala
                        for renglon_sala in salas:
                            hoja.cell(row=row+1, column=1).value = imprimir_sala
                               
                for numero_cliente, cliente_nombre, cliente_apellido in clientes:                    
                    if numero_cliente == num_cliente_buscar:
                        imprimir_cliente=cliente_nombre
                        imprimir_apellido=cliente_apellido
                        for renglon_clientes in clientes:
                            hoja.cell(row=row+1, column=2).value = imprimir_cliente
                            hoja.cell(row=row+1, column=3).value = imprimir_apellido
                              
                for num_cliente, numero_evento, evento_nombre, clave_sala, evento_turno, evento_fecha in eventos:                                    
                    if numero_evento==id_buscar:
                        imprimir_evento=evento_nombre
                        imprimir_turno=evento_turno
                        for renglon_ev_turn in eventos:
                            hoja.cell(row=row+1, column=4).value = imprimir_evento
                            hoja.cell(row=row+1, column=5).value = imprimir_turno
                            
                print("--------------------------------------------------------------------")
                print(f"{imprimir_sala}\t{imprimir_cliente} {imprimir_apellido}\t\t{imprimir_evento}\t\t{imprimir_turno}")
        print("----------------------------FIN DEL REPORTE----------------------------")
        fechaExistente = True
        
        libro.save('Consulta_eventos_prueba.xlsx')
        
    if opcion == 4:
        
        print("Registrar un nuevo cliente\n")
        while True:
            nombre_cliente=input("Ingrese el nombre del cliente: ")
            if nombre_cliente == "":
                print("El nombre del cliente no puede omitirse\n")
            else:
                apellidos=input("Ingrese los apellidos del cliente: ")
                clave_cliente += 1
                print("Cliente agregado.\n")
                clientes.append((clave_cliente, nombre_cliente, apellidos))
        
                if path.isfile('clientes.csv'):
                    print('El archivo ha sido actualizado\n')
                    
                else:
                    print('El archivo de los clientes no existe, pero ya fue creado con éxito\n')
                
                    archivo_clientes=open("clientes.csv","w", newline="")
                    grabador = csv.writer(archivo_clientes)
                    grabador.writerow(("Clave", "Nombre", "Apellido"))
                    archivo_clientes.close()
                break
    
    if opcion == 5:
        print("Registrar una sala\n")
        while True:
            nombre_sala = input("Ingrese el nombre de la sala: ")
            if nombre_sala == "":
                print("El nombre de la sala no debe omitirse\n")
            else:
                cupo_sala = int(input("Ingrese el cupo de la sala: "))
                if cupo_sala <= 0:
                    print("El cupo de la sala debe ser un numero mayor a 0\n")
                else:
                    clave_sala += 1
                    print("Sala agregada.\n")
                    
                    salas.append((clave_sala, nombre_sala, cupo_sala))
                    
                    if path.isfile('salas.csv'):
                        print('El archivo a sido actualizado\n')
                
                    else:
                        print('El archivo de las salas no existe, pero ya fue creado con éxito\n')
                    
                        archivo_salas=open("salas.csv","w", newline="")
                        grabador = csv.writer(archivo_salas)
                        grabador.writerow(("Clave", "Sala", "Cupo"))
                        archivo_salas.close()
                    break
                
    if opcion == 6:
        print('Salas disponibles\n')
        for clave_sala, nombre_sala, cupo_sala in salas:
            salita=nombre_sala
            total_salas_turnos.append((salita, "Matutino"))
            total_salas_turnos.append((salita, "Vespertino"))
            total_salas_turnos.append((salita, "Nocturno"))            
            conjunto_total_salas=set(total_salas_turnos)
            conjunto_salas_ocupadas=set(salas_turnos_ocupados)                     
            conjunto_salas_disponibles=conjunto_total_salas - conjunto_salas_ocupadas            
        print(conjunto_salas_disponibles)
    
    if opcion == 7:
        print("Usted a salido con éxito\n")
        archivo_clientes=open("clientes.csv","a", newline="")
        grabador = csv.writer(archivo_clientes)
        grabador.writerows(clientes)
        archivo_clientes.close()
        archivo_salas=open("salas.csv","a", newline="")
        grabador = csv.writer(archivo_salas)
        grabador.writerows(salas)
        archivo_salas.close()
        break
    
    
    
    
    
    
    
    
    
    